import os
import shutil
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
from zipfile import ZipFile
import zipfile
import tempfile
import sys
import xlwings as xw
from openpyxl.utils.cell import column_index_from_string
import ctypes
import datetime

sheet_name = None
room_entries = {}

def safe_print(msg):
    try:
        sys.stdout.buffer.write((str(msg) + '\n').encode('utf-8', errors='replace'))
    except Exception:
        pass

def require_admin():
    if not ctypes.windll.shell32.IsUserAnAdmin():
        safe_print("[INFO] ขอสิทธิ์แอดมินใหม่")
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, ' '.join(sys.argv), None, 1)
        sys.exit()

def force_excel_recalculate(file_path: Path):
    safe_print("[INFO] เปิด Excel เพื่อคำนวณสูตร + ลิงก์...")
    app = xw.App(visible=False)
    wb = app.books.open(str(file_path), update_links=True)
    wb.app.api.CalculateFullRebuild()
    wb.api.RefreshAll()
    wb.save()
    wb.close()
    app.quit()
    safe_print("[INFO] คำนวณเสร็จแล้ว และบันทึกไฟล์ใหม่")

def fix_corrupt_drawings(file_path: Path, new_filename: str = None) -> Path:
    with tempfile.TemporaryDirectory() as tmpdir:
        temp_path = Path(tmpdir) / "cleaned.xlsx"
        with zipfile.ZipFile(file_path, 'r') as zin:
            with zipfile.ZipFile(temp_path, 'w') as zout:
                for item in zin.infolist():
                    if "xl/drawings/" in item.filename and "NULL" in item.filename:
                        continue
                    if item.filename.endswith(".rels"):
                        try:
                            data = zin.read(item.filename).decode("utf-8", errors="ignore")
                            if "NULL" in data or 'Target="NULL"' in data:
                                safe_print(f"[INFO] ข้ามไฟล์ .rels ที่เสีย: {item.filename}")
                                continue
                            data = data.encode("utf-8")
                            zout.writestr(item, data)
                            continue
                        except:
                            continue
                    try:
                        data = zin.read(item.filename)
                        zout.writestr(item, data)
                    except:
                        continue

        cleaned = file_path.with_name(f"{new_filename}.xlsx") if new_filename else file_path.with_name(f"CLEANED_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        try:
            shutil.copy(temp_path, cleaned)
        except PermissionError:
            safe_print("[ERROR] เขียนไฟล์ไม่ได้ อาจเปิดอยู่หรือไม่มีสิทธิ์: " + str(cleaned))
            return None
        return cleaned

def unmerge_specific_cells(ws, targets):
    to_unmerge = []
    for rng in ws.merged_cells.ranges:
        for cell in targets:
            if cell in rng:
                to_unmerge.append(rng.coord)
                break
    for coord in to_unmerge:
        safe_print(f"[INFO] ยกเลิก merge: {coord}")
        ws.unmerge_cells(coord)

def write_cell(ws, col_label: str, row: int, value):
    try:
        col = column_index_from_string(col_label)
        ws.cell(row=row, column=col).value = value
    except Exception as e:
        safe_print(f"[WARN] เขียน cell {col_label}{row} ไม่ได้: {e}")

def read_cell(ws, col_label: str, row: int, fallback=None):
    try:
        col = column_index_from_string(col_label)
        val = ws.cell(row=row, column=col).value
        return val if val is not None else fallback
    except Exception as e:
        safe_print(f"[WARN] อ่าน cell {col_label}{row} ไม่ได้: {e}")
        return fallback

def safe_float(val, fallback=0.0):
    try:
        return float(val)
    except:
        return fallback

def update_excel(file_path: Path, water_val, elec_val, save_as_name=None):
    if not file_path.exists():
        safe_print("[ERROR] ไม่พบไฟล์: " + str(file_path))
        return

    if not str(water_val).strip() and not str(elec_val).strip():
        safe_print(f"[SKIP] ไม่พบข้อมูลน้ำและไฟในไฟล์: {file_path.name}")
        return

    # STEP 1: Fix corrupt drawings first
    fixed_file = fix_corrupt_drawings(file_path, new_filename=save_as_name)
    if not fixed_file:
        return

    # STEP 2: Force recalculate AFTER fixing
    try:
        force_excel_recalculate(fixed_file)
    except Exception as e:
        safe_print(f"[WARN] ไม่สามารถคำนวณสูตรใน Excel ได้: {e}")

    # STEP 3: Now open with openpyxl
    try:
        wb = load_workbook(fixed_file, keep_links=False, data_only=False)
        ws = wb.worksheets[0]

        unmerge_specific_cells(ws, ["W2", "T2", "X2", "U2"])

        prev_water = read_cell(ws, "X", 2, fallback=0)
        prev_elec = read_cell(ws, "U", 2, fallback=0)
        safe_print(f"ก่อน :{prev_water}, {prev_elec}")
        safe_print(f"ใหม่ :{water_val}, {elec_val}")
        # Update the values in the specified cells
        write_cell(ws, "W", 2, prev_water)
        write_cell(ws, "T", 2, prev_elec)
        write_cell(ws, "X", 2, safe_float(water_val))
        write_cell(ws, "U", 2, safe_float(elec_val))
        # Update the same values in row 4
        write_cell(ws, "W", 4, prev_water)
        write_cell(ws, "T", 4, prev_elec)
        write_cell(ws, "X", 4, safe_float(water_val))
        write_cell(ws, "U", 4, safe_float(elec_val))
        # Update entry day 
        write_cell(ws,"G",2, entry_days.get().strip())
        write_cell(ws, "G", 4, entry_days.get().strip())
        write_cell(ws, "H", 2, entry_month.get().strip())
        write_cell(ws, "H", 4, entry_month.get().strip())
        write_cell(ws, "I", 2, entry_years.get().strip())
        write_cell(ws, "I", 4, entry_years.get().strip())
        # Update entry day
        write_cell(ws, "V", 2, entry_day.get().strip())
        write_cell(ws, "V", 4, entry_day.get().strip())
        wb.save(fixed_file)
        safe_print("[SUCCESS] อัปเดตไฟล์แล้ว: " + fixed_file.name)

    except Exception as e:
        safe_print("[ERROR] อัปเดตไฟล์ล้มเหลว: " + str(e))

# ขอสิทธิ์ admin
require_admin()

# GUI
class MeterEntry:
    def __init__(self, parent, room_name, row):
        self.room_name = room_name
        tk.Label(parent, text=room_name).grid(row=row, column=0, padx=5, pady=2)
        self.water_entry = tk.Entry(parent, width=10)
        self.water_entry.grid(row=row, column=1, padx=5)
        self.electric_entry = tk.Entry(parent, width=10)
        self.electric_entry.grid(row=row, column=2, padx=5)

    def get_values(self):
        return self.water_entry.get().strip(), self.electric_entry.get().strip()

def browse_folder():
    folder = filedialog.askdirectory()
    if folder:
        entry_path.delete(0, tk.END)
        entry_path.insert(0, folder)
        update_room_list()

def update_room_list():
    for widget in scroll_frame.winfo_children():
        widget.destroy()
    room_entries.clear()

    dorm_root = Path(entry_path.get().strip())
    if not dorm_root.exists():
        return

    row = 1
    tk.Label(scroll_frame, text="ห้อง").grid(row=0, column=0)
    tk.Label(scroll_frame, text="มิเตอร์น้ำ").grid(row=0, column=1)
    tk.Label(scroll_frame, text="มิเตอร์ไฟ").grid(row=0, column=2)

    for floor in dorm_root.iterdir():
        if floor.is_dir():
            for room in floor.iterdir():
                if room.is_dir() and room.name.startswith("ห้อง"):
                    entry = MeterEntry(scroll_frame, room.name, row)
                    room_entries[room] = entry
                    row += 1

def run_process():
    year = entry_year.get().strip()
    old_file_name = entry_old_file.get().strip()
    new_file_name = entry_new_file.get().strip()

    for room_path, meter_entry in room_entries.items():
        file_path = room_path / f"ปี {year}" / f"{new_file_name}.xlsx"
        old_file = room_path / f"ปี {year}" / f"{old_file_name}.xlsx"
        if not old_file.exists():
            safe_print(f"[SKIP] ไม่พบไฟล์เก่า: {old_file}")
            continue

        water_val, elec_val = meter_entry.get_values()
        if not water_val and not elec_val:
            safe_print(f"[SKIP] ไม่มีข้อมูลน้ำและไฟในห้อง: {room_path.name}")
            continue

        shutil.copy(old_file, file_path)
        try:
            update_excel(file_path, water_val, elec_val, save_as_name=new_file_name)
        except Exception as e:
            safe_print(f"[ERROR] {room_path}: {e}")

root = tk.Tk()
root.title("ระบบบันทึกค่าน้ำค่าไฟตามห้อง")

tk.Label(root, text="📁 โฟลเดอร์หลัก:").grid(row=0, column=0, sticky="e")
entry_path = tk.Entry(root, width=40)
entry_path.grid(row=0, column=1)
tk.Button(root, text="เลือก", command=browse_folder).grid(row=0, column=2)

tk.Label(root, text="📅 ปี:").grid(row=1, column=0, sticky="e")
entry_year = tk.Entry(root)
entry_year.grid(row=1, column=1)

tk.Label(root, text="📁 ไฟล์เก่า:").grid(row=2, column=0, sticky="e")
entry_old_file = tk.Entry(root)
entry_old_file.grid(row=2, column=1)

tk.Label(root, text="🆕 ไฟล์ใหม่:").grid(row=3, column=0, sticky="e")
entry_new_file = tk.Entry(root)
entry_new_file.grid(row=3, column=1)

tk.Label(root, text="🆕 วันที่จด:").grid(row=4, column=0, sticky="e")
entry_day = tk.Entry(root)
entry_day.grid(row=4, column=1)

row_frame = tk.Frame(root)
row_frame.grid(row=5, column=0, columnspan=6, sticky="w")

tk.Label(row_frame, text="🗓️ วันที่แจก:").pack(side="left")
entry_days = tk.Entry(row_frame, width=6)
entry_days.pack(side="left")

tk.Label(row_frame, text="เดือน:").pack(side="left", padx=(10, 0))  
entry_month = tk.Entry(row_frame, width=6)
entry_month.pack(side="left")

tk.Label(row_frame, text="ปี:").pack(side="left", padx=(10, 0))
entry_years = tk.Entry(row_frame, width=6)
entry_years.pack(side="left")

canvas = tk.Canvas(root, height=400)
frame_rooms = tk.Frame(canvas)
scrollbar = tk.Scrollbar(root, orient="vertical", command=canvas.yview)
canvas.configure(yscrollcommand=scrollbar.set)

scrollbar.grid(row=6, column=3, sticky="ns")
canvas.grid(row=6, column=0, columnspan=3)
canvas.create_window((0, 0), window=frame_rooms, anchor="nw")

frame_rooms.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
scroll_frame = frame_rooms

room_entries = {}
tk.Button(root, text="✅ อัปเดตทั้งหมด", command=run_process, bg="lightgreen").grid(row=7, column=0, columnspan=3, pady=10)

root.mainloop()
